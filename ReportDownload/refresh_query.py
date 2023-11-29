import os
import pyxlsb
from openpyxl import load_workbook

# 엑셀 파일 경로
input_file_path = "Z:/02.펀드/003.매매보고서 대사/KIS-PBS_스왑매매내역.xlsx"

# 엑셀 파일을 열고 Power Query 새로고침
def refresh_power_query(input_file_path):
    # 엑셀 파일 불러오기
    workbook = load_workbook(filename=input_file_path, data_only=True)

    # 모든 시트에 대해 Power Query 새로고침 수행
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        try:
            # .xlsx 파일은 QueryTables 메서드를 사용하여 Power Query 새로고침 가능
            sheet.QueryTables(1).Refresh()
            print(f"Power Query in sheet '{sheet_name}' refreshed successfully.")
        except Exception as e:
            print(f"Error refreshing Power Query in sheet '{sheet_name}': {str(e)}")

    # 업데이트된 내용을 저장
    workbook.save(input_file_path)

# Power Query 새로고침 함수 호출
refresh_power_query(input_file_path)